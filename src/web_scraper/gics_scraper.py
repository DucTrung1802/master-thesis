# src\web_scraper\gics_scraper.py

# ===== Standard Library =====
import csv
import os
import re
import time
from typing import Dict, List

# ===== Third-Party Libraries =====
import requests
import openpyxl

# ===== Local / Custom Modules =====
from logger.logger import Logger
from utils.constants import (
    GICS_RAW_DATA_DIR,
    SCRAPER_RETRY_ATTEMPTS,
    SCRAPER_RETRY_DELAY,
)
from web_scraper.base_scraper import BaseScraper, register_scraper


@register_scraper
class GicsScraper(BaseScraper):
    """Download and parse the official MSCI GICS structure (the global industry
    classification taxonomy) into a flat reference CSV.

    Source: MSCI's published Excel "GICS structure & definitions effective close
    of 17 March 2023" — the newest GICS revision: 11 sectors / 25 industry groups
    / 74 industries / 163 sub-industries, each identified by an 8-digit code that
    is self-describing (sector = digits 1-2, industry group = 1-4, industry = 1-6,
    sub-industry = all 8).

    The "GICS structure eff March 2023" sheet is a *pro-forma* markup of the 2023
    change, so it needs cleaning:
      - codes appear only on the first occurrence of each level (forward-fill);
      - it carries 7 extra "(Discontinued)" sub-industry rows (170 -> drop to 163)
        and 2 discontinued industries (76 -> 74);
      - names carry change annotations like "(New Code)" / "(Definition Update)"
        that are stripped, while semantic ones like "(HMOs)" / "(except bauxite)"
        are kept;
      - each sub-industry's definition sits in the row directly below it.

    Output: GICS_RAW_DATA_DIR/gics_2023_official.csv — one row per sub-industry,
    all four levels with code + name + snake_case, plus the sub-industry definition.
    The raw .xlsx is also kept under GICS_RAW_DATA_DIR for provenance.
    """

    SOURCE_NAME = "gics"

    # Full MSCI document URL (the GUID path segment + ?t= token are required to
    # resolve the file).
    STRUCTURE_URL = (
        "https://www.msci.com/documents/1296102/29559863/"
        "GICS_structure_and_definitions_effective_close_of_March_17_2023.xlsx/"
        "e47b8086-56fd-c9d2-196f-c2054b24b1d4?t=1670964718735"
    )
    SHEET_NAME = "GICS structure eff March 2023"
    XLSX_FILENAME = "gics_structure_eff_2023-03-17.xlsx"
    OUTPUT_FILENAME = "gics_2023_official.csv"

    # Expected level counts for the 2023 revision — used only as a sanity check.
    EXPECTED_COUNTS = (11, 25, 74, 163)  # sectors, groups, industries, sub-industries

    OUTPUT_COLUMNS = [
        "sector_code", "sector", "sector_snake",
        "industry_group_code", "industry_group", "industry_group_snake",
        "industry_code", "industry", "industry_snake",
        "sub_industry_code", "sub_industry", "sub_industry_snake",
        "sub_industry_definition",
    ]

    # A parenthetical in a name is dropped if it contains any of these words
    # (it marks a 2023 change); all other parentheticals are kept as semantic.
    _CHANGE_WORDS = ("new", "discontinued", "definition update", "sector change", "code")

    def __init__(
        self,
        logger: Logger,
        switch_handler=None,
        power: int = 10,
        retry_attempts: int = SCRAPER_RETRY_ATTEMPTS,
        retry_delay: float = SCRAPER_RETRY_DELAY,
    ):
        super().__init__(
            logger=logger,
            switch_handler=switch_handler,
            power=power,
            retry_attempts=retry_attempts,
            retry_delay=retry_delay,
        )
        self._session = requests.Session()
        self._session.headers.update({
            "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/123.0.0.0 Safari/537.36"),
        })

    # ──────────────────────────────────────────────────────────────────────
    # Download
    # ──────────────────────────────────────────────────────────────────────

    def _download_structure(self, dest_path: str) -> bool:
        """Download the MSCI GICS structure xlsx to dest_path (retried).

        Written to a temp file then atomically renamed, so an interrupted run
        never leaves a truncated xlsx behind. Returns True on success.
        """
        for attempt in range(1, self._retry_attempts + 1):
            try:
                r = self._session.get(self.STRUCTURE_URL, timeout=60)
                r.raise_for_status()
                tmp_path = dest_path + ".tmp"
                with open(tmp_path, "wb") as f:
                    f.write(r.content)
                os.replace(tmp_path, dest_path)
                self._logger.log_info(
                    f"GICS: downloaded structure ({len(r.content)} bytes) -> {dest_path}"
                )
                return True
            except Exception as e:
                self._logger.log_warning(
                    f"GICS: download attempt {attempt}/{self._retry_attempts} failed: {e}"
                )
                if attempt < self._retry_attempts:
                    time.sleep(self._retry_delay)
        self._logger.log_error("GICS: failed to download structure xlsx.")
        return False

    # ──────────────────────────────────────────────────────────────────────
    # Parsing helpers
    # ──────────────────────────────────────────────────────────────────────

    @staticmethod
    def _code(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            v = int(v)
        return str(v).strip()

    @staticmethod
    def _txt(v) -> str:
        return "" if v is None else str(v).replace("\n", " ").strip()

    @staticmethod
    def _is_discontinued(name: str) -> bool:
        return "discontinued" in name.lower()

    def _clean_name(self, name: str) -> str:
        """Strip change-annotation parentheticals (e.g. '(New Code)') while keeping
        semantic ones (e.g. '(HMOs)', '(except bauxite)')."""
        def repl(match: "re.Match") -> str:
            inside = match.group(0).lower()
            return "" if any(w in inside for w in self._CHANGE_WORDS) else match.group(0)

        name = re.sub(r"\([^)]*\)", repl, name)
        return re.sub(r"\s{2,}", " ", name).strip(" ,/")

    @staticmethod
    def _snake(name: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", name.replace("&", "and").lower()).strip("_")

    def _parse_structure(self, xlsx_path: str) -> List[Dict]:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if self.SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"GICS: sheet '{self.SHEET_NAME}' not found in {xlsx_path}; "
                f"sheets present: {wb.sheetnames}"
            )
        ws = wb[self.SHEET_NAME]

        rows: List[Dict] = []
        cur = {"sc": "", "sn": "", "gc": "", "gn": "", "ic": "", "in": ""}
        last: Dict = None

        # Data starts at row 6 (rows 1-5 are the title, colour key, and header).
        # Columns: A/B sector, C/D industry group, E/F industry, G/H sub-industry.
        for r in range(6, ws.max_row + 1):
            sec_c, sec_n = self._code(ws.cell(r, 1).value), self._txt(ws.cell(r, 2).value)
            grp_c, grp_n = self._code(ws.cell(r, 3).value), self._txt(ws.cell(r, 4).value)
            ind_c, ind_n = self._code(ws.cell(r, 5).value), self._txt(ws.cell(r, 6).value)
            sub_c, sub_n = self._code(ws.cell(r, 7).value), self._txt(ws.cell(r, 8).value)

            # Forward-fill the higher levels; ignore discontinued ones so they
            # never get attached to a kept sub-industry.
            if sec_c and not self._is_discontinued(sec_n):
                cur["sc"], cur["sn"] = sec_c, self._clean_name(sec_n)
            if grp_c and not self._is_discontinued(grp_n):
                cur["gc"], cur["gn"] = grp_c, self._clean_name(grp_n)
            if ind_c and not self._is_discontinued(ind_n):
                cur["ic"], cur["in"] = ind_c, self._clean_name(ind_n)

            if re.fullmatch(r"\d{8}", sub_c):
                if self._is_discontinued(sub_n):
                    last = None  # drop it, and skip its definition line(s)
                    continue
                last = {
                    "sector_code": cur["sc"], "sector": cur["sn"],
                    "industry_group_code": cur["gc"], "industry_group": cur["gn"],
                    "industry_code": cur["ic"], "industry": cur["in"],
                    "sub_industry_code": sub_c,
                    "sub_industry": self._clean_name(sub_n),
                    "sub_industry_definition": "",
                }
                rows.append(last)
            elif last is not None and not sub_c and sub_n:
                # Definition line(s) sit directly under each sub-industry row.
                last["sub_industry_definition"] = (
                    last["sub_industry_definition"] + " " + sub_n
                ).strip()

        for row in rows:
            row["sector_snake"] = self._snake(row["sector"])
            row["industry_group_snake"] = self._snake(row["industry_group"])
            row["industry_snake"] = self._snake(row["industry"])
            row["sub_industry_snake"] = self._snake(row["sub_industry"])

        self._validate(rows)
        return rows

    def _validate(self, rows: List[Dict]) -> None:
        counts = (
            len({r["sector_code"] for r in rows}),
            len({r["industry_group_code"] for r in rows}),
            len({r["industry_code"] for r in rows}),
            len({r["sub_industry_code"] for r in rows}),
        )
        if counts == self.EXPECTED_COUNTS:
            self._logger.log_info(
                f"GICS: parsed {counts[0]} sectors / {counts[1]} groups / "
                f"{counts[2]} industries / {counts[3]} sub-industries (matches 2023)."
            )
        else:
            self._logger.log_warning(
                f"GICS: parsed counts {counts} differ from expected "
                f"{self.EXPECTED_COUNTS}; MSCI may have revised the structure."
            )

    # ──────────────────────────────────────────────────────────────────────
    # Output
    # ──────────────────────────────────────────────────────────────────────

    def _write_csv(self, rows: List[Dict], out_path: str) -> None:
        tmp_path = out_path + ".tmp"
        with open(tmp_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=self.OUTPUT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        os.replace(tmp_path, out_path)

    # ──────────────────────────────────────────────────────────────────────
    # Entry point
    # ──────────────────────────────────────────────────────────────────────

    def scrape(self) -> None:
        """BaseScraper entry point: download the MSCI GICS structure and write the
        cleaned taxonomy CSV to GICS_RAW_DATA_DIR/gics_2023_official.csv."""
        os.makedirs(GICS_RAW_DATA_DIR, exist_ok=True)
        xlsx_path = os.path.join(GICS_RAW_DATA_DIR, self.XLSX_FILENAME)
        out_path = os.path.join(GICS_RAW_DATA_DIR, self.OUTPUT_FILENAME)

        self._logger.log_info("GICS: scraping official GICS structure from MSCI.")
        if not self._download_structure(xlsx_path):
            self._logger.log_error("GICS: aborting — could not download structure.")
            return

        rows = self._parse_structure(xlsx_path)
        self._write_csv(rows, out_path)
        self._logger.log_info(
            f"GICS: saved {len(rows)} sub-industries -> {out_path}"
        )
